import io
import uuid
from datetime import datetime, timezone
from typing import Dict, Any, List, Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import and_

from app.modules.inventory.models import InventoryItem
from app.modules.scoring.service import calculate_item_health_score
from app.modules.report.schemas import ReportFilterRequest, ReportSummaryItem

# ReportLab libraries
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# Openpyxl libraries
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ReportService:
    @staticmethod
    async def generate_report_data(db: AsyncSession, filters: ReportFilterRequest) -> Dict[str, Any]:
        # 1. Construct SQL Query
        query = select(InventoryItem)
        conditions = []

        if filters.category and filters.category != "All":
            conditions.append(InventoryItem.category == filters.category)
        
        if filters.storage_location and filters.storage_location != "All":
            conditions.append(InventoryItem.storage_location == filters.storage_location)

        if filters.start_date:
            try:
                start_dt = datetime.fromisoformat(filters.start_date.replace("Z", "+00:00"))
                conditions.append(InventoryItem.entry_date >= start_dt)
            except ValueError:
                pass

        if filters.end_date:
            try:
                end_dt = datetime.fromisoformat(filters.end_date.replace("Z", "+00:00"))
                conditions.append(InventoryItem.entry_date <= end_dt)
            except ValueError:
                pass

        if conditions:
            query = query.where(and_(*conditions))

        result = await db.execute(query)
        items = result.scalars().all()

        # 2. Compute dynamic sub-scores and construct details list
        summary_items: List[ReportSummaryItem] = []
        total_freshness = 0.0
        freshness_counts = {"Fresh": 0, "Good": 0, "Acceptable": 0, "Near Spoilage": 0, "Spoiled": 0}
        
        total_remaining_shelf_life = 0.0
        shelf_life_warnings = 0
        fefo_overruns = 0
        
        total_visual_score = 0.0
        mold_count = 0
        
        potential_waste_qty = 0.0
        use_soon_recommendations = 0

        total_temp_deviation = 0.0
        total_hum_deviation = 0.0
        compliant_storage_count = 0

        for item in items:
            health = await calculate_item_health_score(db, item)
            score = health.combined_health_score
            classification = health.quality_classification
            
            # Shelf life predictions lookup
            remaining_shelf_life = health.breakdown.shelflife_score * 0.1 # Estimate remaining days from score ratio
            
            # Map classification count
            if score >= 85.0:
                freshness_counts["Fresh"] += 1
            elif score >= 70.0:
                freshness_counts["Good"] += 1
            elif score >= 50.0:
                freshness_counts["Acceptable"] += 1
            elif score >= 30.0:
                freshness_counts["Near Spoilage"] += 1
            else:
                freshness_counts["Spoiled"] += 1

            # Accumulate metrics
            total_freshness += score
            total_remaining_shelf_life += remaining_shelf_life
            
            # Expiry timeline checks
            days_left = (item.expiry_date.replace(tzinfo=timezone.utc) - datetime.now(timezone.utc)).days
            if days_left <= 3:
                shelf_life_warnings += 1
                use_soon_recommendations += 1

            if remaining_shelf_life < days_left - 1.5:
                fefo_overruns += 1

            # Visual & Spoilage attributes
            total_visual_score += health.breakdown.visual_score
            if score < 50.0:
                mold_count += 1
                potential_waste_qty += item.quantity

            # Storage compliance
            if health.breakdown.storage_score >= 80.0:
                compliant_storage_count += 1

            summary_items.append(
                ReportSummaryItem(
                    id=str(item.id),
                    name=item.name,
                    category=item.category,
                    storage_location=item.storage_location or "N/A",
                    entry_date=item.entry_date.isoformat(),
                    expiry_date=item.expiry_date.isoformat(),
                    quantity=item.quantity,
                    unit=item.unit,
                    freshness_score=score,
                    remaining_shelf_life_days=max(0.0, float(round(remaining_shelf_life, 1))),
                    status=classification
                )
            )

        item_count = len(items)
        avg_freshness = total_freshness / item_count if item_count > 0 else 100.0
        avg_remaining_days = total_remaining_shelf_life / item_count if item_count > 0 else 0.0
        avg_visual = total_visual_score / item_count if item_count > 0 else 100.0
        storage_compliance_rate = (compliant_storage_count / item_count * 100.0) if item_count > 0 else 100.0

        # Construct specific stats block depending on report type
        stats: Dict[str, Any] = {}
        if filters.report_type == "freshness":
            stats = {
                "total_items": item_count,
                "average_freshness_score": round(avg_freshness, 1),
                "distribution": freshness_counts,
                "fresh_percentage": round((freshness_counts["Fresh"] + freshness_counts["Good"]) / item_count * 100.0, 1) if item_count > 0 else 100.0
            }
        elif filters.report_type == "shelf-life":
            stats = {
                "total_items": item_count,
                "average_remaining_days": round(avg_remaining_days, 1),
                "shelf_life_warnings": shelf_life_warnings,
                "fefo_overruns": fefo_overruns
            }
        elif filters.report_type == "quality":
            stats = {
                "total_items": item_count,
                "average_visual_condition": round(avg_visual, 1),
                "potential_mold_alerts": mold_count,
                "critical_quality_warnings": freshness_counts["Near Spoilage"] + freshness_counts["Spoiled"]
            }
        elif filters.report_type == "waste":
            stats = {
                "total_items": item_count,
                "potential_waste_qty": round(potential_waste_qty, 1),
                "use_soon_count": use_soon_recommendations,
                "risk_items_count": freshness_counts["Near Spoilage"] + freshness_counts["Spoiled"]
            }
        elif filters.report_type == "storage":
            stats = {
                "total_items": item_count,
                "storage_compliance_rate": round(storage_compliance_rate, 1),
                "non_compliant_count": item_count - compliant_storage_count,
                "ideal_rate_achieved": storage_compliance_rate >= 90.0
            }

        return {
            "report_type": filters.report_type,
            "generated_at": datetime.now(timezone.utc),
            "summary_stats": stats,
            "items": summary_items
        }

    @staticmethod
    def generate_pdf(report_type: str, data: Dict[str, Any]) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()
        
        # Define clean, modern color accents (Forest Green & Emerald shades)
        c_primary = colors.HexColor("#064E3B") # Deep forest green
        c_secondary = colors.HexColor("#10B981") # Brand Emerald
        c_text = colors.HexColor("#1E293B") # Slate dark
        c_border = colors.HexColor("#E2E8F0") # Slate light border

        # Custom Paragraph styles
        title_style = ParagraphStyle(
            "DocTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=24,
            textColor=c_primary,
            alignment=0,
            spaceAfter=15
        )
        
        h2_style = ParagraphStyle(
            "DocH2",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=14,
            textColor=c_primary,
            spaceBefore=12,
            spaceAfter=6
        )

        body_style = ParagraphStyle(
            "DocBody",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=10,
            textColor=c_text,
            spaceAfter=6
        )

        bold_style = ParagraphStyle(
            "DocBold",
            parent=body_style,
            fontName="Helvetica-Bold"
        )

        table_header_style = ParagraphStyle(
            "TableHeader",
            parent=body_style,
            fontName="Helvetica-Bold",
            textColor=colors.white
        )

        story = []

        # 1. Header Section
        story.append(Paragraph(f"FreshLens Analytics Hub", title_style))
        story.append(Paragraph(f"Report Type: <b>{report_type.upper()} REPORT</b>", body_style))
        story.append(Paragraph(f"Generated At: {data['generated_at'].strftime('%Y-%m-%d %H:%M:%S UTC')}", body_style))
        story.append(Spacer(1, 15))

        # 2. Summary Card Section
        story.append(Paragraph("Executive Summary Stats", h2_style))
        summary_data = []
        for key, value in data["summary_stats"].items():
            key_formatted = key.replace("_", " ").title()
            if isinstance(value, dict):
                val_formatted = ", ".join(f"{k}: {v}" for k, v in value.items())
            else:
                val_formatted = str(value)
            summary_data.append([Paragraph(f"<b>{key_formatted}</b>", body_style), Paragraph(val_formatted, body_style)])

        summary_table = Table(summary_data, colWidths=[200, 340])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
            ('BOX', (0, 0), (-1, -1), 1, c_border),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, c_border),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))

        # 3. Details Table
        story.append(Paragraph("Inventory Item Breakdown Details", h2_style))
        
        # Col Widths total: 540 (Letter page width 612 - margins 72)
        # Widths: Name(140), Category(90), Storage(95), Qty(60), Freshness(75), Status(80)
        table_headers = [
            Paragraph("Item Name", table_header_style),
            Paragraph("Category", table_header_style),
            Paragraph("Location", table_header_style),
            Paragraph("Qty", table_header_style),
            Paragraph("Freshness", table_header_style),
            Paragraph("Status", table_header_style)
        ]
        
        detail_rows = [table_headers]
        for idx, item in enumerate(data["items"]):
            # Set status indicator color dynamically
            status_text = item.status
            if status_text in ("Near Spoilage", "Spoiled", "SPOILED"):
                status_html = f"<font color='#EF4444'><b>{status_text}</b></font>"
            elif status_text in ("Warning", "WARNING", "Acceptable"):
                status_html = f"<font color='#F59E0B'><b>{status_text}</b></font>"
            else:
                status_html = f"<font color='#10B981'><b>{status_text}</b></font>"

            detail_rows.append([
                Paragraph(item.name, body_style),
                Paragraph(item.category, body_style),
                Paragraph(item.storage_location, body_style),
                Paragraph(f"{item.quantity} {item.unit}", body_style),
                Paragraph(f"{item.freshness_score}%", body_style),
                Paragraph(status_html, body_style)
            ])

        details_table = Table(detail_rows, colWidths=[140, 90, 95, 60, 75, 80])
        
        # Build TableStyle
        ts = [
            ('BACKGROUND', (0, 0), (-1, 0), c_primary),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, c_border),
            ('BOX', (0, 0), (-1, -1), 1, c_border),
        ]
        # Add zebra striping
        for i in range(1, len(detail_rows)):
            if i % 2 == 0:
                ts.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F8FAFC")))
                
        details_table.setStyle(TableStyle(ts))
        story.append(details_table)

        # Build PDF Document
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_excel(report_type: str, data: Dict[str, Any]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary Report"

        # Define Colors & Fonts
        fill_header = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_title = Font(name="Calibri", size=16, bold=True, color="064E3B")
        font_h2 = Font(name="Calibri", size=12, bold=True, color="000000")
        font_bold = Font(name="Calibri", size=11, bold=True)
        
        border_thin = Border(
            left=Side(style='thin', color="E2E8F0"),
            right=Side(style='thin', color="E2E8F0"),
            top=Side(style='thin', color="E2E8F0"),
            bottom=Side(style='thin', color="E2E8F0")
        )

        # 1. Document Title
        ws.append(["FreshLens Inventory Analytics Report"])
        ws["A1"].font = font_title
        ws.row_dimensions[1].height = 30
        
        ws.append([f"Report Type: {report_type.upper()}"])
        ws.append([f"Generated At: {data['generated_at'].strftime('%Y-%m-%d %H:%M:%S UTC')}"])
        ws.append([]) # spacer

        # 2. Summary stats
        ws.append(["Executive Summary"])
        ws.cell(row=ws.max_row, column=1).font = font_h2
        
        for key, val in data["summary_stats"].items():
            key_formatted = key.replace("_", " ").title()
            if isinstance(val, dict):
                val_formatted = ", ".join(f"{k}: {v}" for k, v in val.items())
            else:
                val_formatted = val
            ws.append([key_formatted, val_formatted])
            ws.cell(row=ws.max_row, column=1).font = font_bold
            ws.cell(row=ws.max_row, column=1).border = border_thin
            ws.cell(row=ws.max_row, column=2).border = border_thin

        ws.append([]) # spacer
        ws.append([]) # spacer

        # 3. Details Table
        ws.append(["Inventory Details"])
        ws.cell(row=ws.max_row, column=1).font = font_h2
        
        headers = ["Item Name", "Category", "Storage Location", "Quantity", "Unit", "Freshness Score", "Remaining Shelf Life (Days)", "Status"]
        ws.append(headers)
        
        header_row_idx = ws.max_row
        ws.row_dimensions[header_row_idx].height = 24
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border_thin

        # Write data rows
        start_data_row = ws.max_row + 1
        for idx, item in enumerate(data["items"]):
            ws.append([
                item.name,
                item.category,
                item.storage_location,
                item.quantity,
                item.unit,
                item.freshness_score,
                item.remaining_shelf_life_days,
                item.status
            ])
            curr_row = ws.max_row
            ws.row_dimensions[curr_row].height = 18
            
            # Formatting & Zebra striping
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.border = border_thin
                if idx % 2 == 1:
                    cell.fill = fill_zebra
                    
                # Format numeric columns
                if col_idx in (4, 6, 7):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Autofit Column Widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Save workbook to memory
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()
